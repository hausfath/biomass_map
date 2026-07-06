#!/usr/bin/env python3
"""
NUTS-2 feedstock supply for the EU BiCRS map (~290 regions, EU-27 + UK + Norway).

Method — "ENSPRESO NUTS-2 distribution, scaled to global-tool country totals" (the proven
US-map pattern). Each feedstock's *within-country distribution* comes from JRC ENSPRESO
(NUTS-2 biomass potential, PJ); each region is then scaled so a country's NUTS-2 regions sum
to that country's total already in the global tool (data/processed/feedstocks.json), which is
anchored to literature/Eurostat. Units cancel — only relative shares matter — so the country
totals stay exact and the regional geography is real.

Streams from ENSPRESO (scenario ENS_Med, year 2020), at NUTS-2:
  ag residues  <- MINBIOAGRW1
  forestry     <- MINBIOFRSR1/1a (forest residues) + MINBIOWOOW1/1a (secondary wood residues)
  manure       <- MINBIOGAS1 (manure_liq + manure_sol, biogas feedstock)
MSW (biogenic) & WWTP biosolids are NUTS-0 in ENSPRESO, so they are allocated to NUTS-2 by
population (Eurostat demo_r_pjanaggr3). Purpose-grown energy/biofuel crops (MINBIOCRP*,
MINBIOLIQ*, MINBIORPS*) are EXCLUDED (Frontier exclusions).

NUTS version: ENSPRESO is NUTS v2013; geometry is v2021. Unmatched regions are remapped via
ENSPRESO's own "NUTS2 conversion" sheet (e.g. FRB0=FR24); any still-unmatched region falls
back to population-share allocation of its country total (documented).

Output: data/processed/feedstocks_eu_nuts.json
"""
import json
import os
import re
import sys
from collections import defaultdict

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # scripts/
from engine_core import ODT_TO_CO2

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
RAW = os.path.join(ROOT, "data", "geo", "eu_raw")
PROC = os.path.join(ROOT, "data", "processed")

XLSX = os.path.join(RAW, "ENSPRESO_BIOMASS.xlsx")
POP = os.path.join(RAW, "eurostat_pop.json")        # 2023 (UK absent post-Brexit)
POP_FALLBACK = os.path.join(RAW, "eurostat_pop_2019.json")  # 2019 (covers UK)
NUTS = os.path.join(ROOT, "data", "geo", "eu_nuts.json")
COUNTRY_FEED = os.path.join(PROC, "feedstocks.json")
UK_WEIGHTS = os.path.join(ROOT, "data", "geo", "uk_feedstock_weights.json")  # item 11: UK Census
OUT = os.path.join(PROC, "feedstocks_eu_nuts.json")

SCENARIO, YEAR = "ENS_Med", "2020"
AG_CODES = {"MINBIOAGRW1"}
FOR_CODES = {"MINBIOFRSR1", "MINBIOFRSR1a", "MINBIOWOOW1", "MINBIOWOOW1a"}
MAN_CODES = {"MINBIOGAS1"}

NUTS2_RE = re.compile(r"\b([A-Z]{2}[0-9A-Z]{2})\b")  # 4-char NUTS-2 token


def load_country_totals():
    fe = json.load(open(COUNTRY_FEED))

    def v(r, k):
        x = r.get(k)
        return (x or {}).get("value", 0) if isinstance(x, dict) else 0
    out = {}
    for r in fe:
        if r.get("level") != "country":
            continue
        out[r["id"]] = {
            "ag": v(r, "ag_residues_odt_mt"),
            "forestry": v(r, "forestry_residues_odt_mt"),
            "manure": v(r, "animal_manure_odt_mt"),
            "msw": v(r, "msw_total_mt"),
            "biofrac": v(r, "msw_biogenic_frac") or 0.5,
            "wwtp": v(r, "human_wwtp_odt_mt"),
            "nutrient_status": r.get("nutrient_status", "moderate"),
        }
    return out


def _load_pop_file(path):
    d = json.load(open(path))
    idx = d["dimension"]["geo"]["category"]["index"]   # code -> position
    val = d["value"]                                   # only geo varies -> pos == flat index
    out = {}
    for code, pos in idx.items():
        v = val.get(str(pos))
        if v is not None and len(code) == 4:
            out[code] = float(v)
    return out


def load_population():
    """NUTS-2 population, preferring the recent year; fall back to 2019 (covers UK post-Brexit)."""
    pop = _load_pop_file(POP)
    if os.path.exists(POP_FALLBACK):
        for code, v in _load_pop_file(POP_FALLBACK).items():
            pop.setdefault(code, v)
    return pop


def load_uk_weights():
    """UK ITL-2 manure/ag distribution weights from the UK June Census (item 11), keyed by 2021
    code. Replaces the population/ENSPRESO-imputed weights for UK manure & ag residues so, e.g.,
    London no longer gets phantom manure. Absent file -> {} (falls back to the ENSPRESO path)."""
    if not os.path.exists(UK_WEIGHTS):
        print("  (UK census weights not found — UK falls back to ENSPRESO/population; run "
              "build_uk_ag_weights.py)")
        return {}
    return {k: v for k, v in json.load(open(UK_WEIGHTS)).items() if not k.startswith("_")}


def load_conversion(wb):
    """new(v2021/2016) code -> set(old v2013 NUTS-2 codes), from ENSPRESO's conversion sheet."""
    ws = wb["NUTS2 conversion"]
    rows = list(ws.iter_rows(values_only=True))
    new2old = defaultdict(set)
    for row in rows[1:]:
        old_c = (str(row[0]).strip() if row[0] else "")
        new_c = (str(row[1]).strip() if row[1] else "")
        expl = (str(row[4]).strip() if len(row) > 4 and row[4] else "")
        if expl and "=" in expl:
            lhs, rhs = expl.split("=", 1)
            new = lhs.strip()
            olds = NUTS2_RE.findall(rhs)
            if len(new) == 4:
                for o in olds:
                    new2old[new].add(o)
        elif new_c and old_c and len(new_c) == 4 and len(old_c) == 4:
            new2old[new_c].add(old_c)
    return new2old


def load_enspreso(wb):
    """{nuts2_code: {ag, forestry, manure}} in PJ, scenario/year fixed."""
    ws = wb["ENER - NUTS2 BioCom E"]
    it = ws.iter_rows(values_only=True)
    next(it)
    pj = defaultdict(lambda: {"ag": 0.0, "forestry": 0.0, "manure": 0.0})
    for r in it:
        if r[0] is None:
            continue
        if str(r[0]) != YEAR or str(r[1]) != SCENARIO:
            continue
        n2, ec, val = r[3], r[4], r[6]
        if not n2 or n2 == "-" or val is None:
            continue
        try:
            val = float(val)
        except (TypeError, ValueError):
            continue
        if ec in AG_CODES:
            pj[n2]["ag"] += val
        elif ec in FOR_CODES:
            pj[n2]["forestry"] += val
        elif ec in MAN_CODES:
            pj[n2]["manure"] += val
    return pj


def est(value, source="", notes=""):
    value = round(value, 4)
    return {"value": value, "low": round(value * 0.6, 4), "high": round(value * 1.5, 4),
            "source": source, "notes": notes}


def dominant(ag, forestry, manure, msw_bio):
    co2 = {"dry": (ag + forestry) * ODT_TO_CO2, "manure_wet": manure * ODT_TO_CO2, "msw": msw_bio}
    top = max(co2, key=co2.get)
    if co2[top] <= 0:
        return "mixed"
    if top == "dry":
        return "forestry_woody" if forestry > ag else "ag_dry"
    return top


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    pj = load_enspreso(wb)
    new2old = load_conversion(wb)
    pop = load_population()
    uk_weights = load_uk_weights()
    ctot = load_country_totals()
    regions = [f["properties"] for f in json.load(open(NUTS))["features"]]

    # Resolve each region's ENSPRESO PJ (direct -> conversion remap -> None for pop-fallback).
    region_pj = {}
    fallbacks = []
    for p in regions:
        code = p["nuts_id"]
        if code in pj:
            region_pj[code] = dict(pj[code])
        elif code in new2old and any(o in pj for o in new2old[code]):
            agg = {"ag": 0.0, "forestry": 0.0, "manure": 0.0}
            for o in new2old[code]:
                if o in pj:
                    for k in agg:
                        agg[k] += pj[o][k]
            region_pj[code] = agg
        else:
            region_pj[code] = None
            fallbacks.append(code)

    # Group by country (ISO3).
    by_country = defaultdict(list)
    for p in regions:
        by_country[p["country"]].append(p)

    records = []
    for iso3, regs in by_country.items():
        tot = ctot.get(iso3)
        if not tot:
            continue  # no country anchor (shouldn't happen — all 29 present)

        # --- raw per-region weights (PJ for ag/for/manure; population for msw/wwtp) ---
        # Matched regions use ENSPRESO PJ; ENSPRESO-missing regions impute PJ from the country's
        # matched per-capita rate (pop x matched_PJ/matched_pop). A final per-country, per-stream
        # normalization then rescales every region so the country sum equals the global-tool total
        # EXACTLY — robust regardless of the matched/imputed split.
        sum_matched_pj = {"ag": 0.0, "forestry": 0.0, "manure": 0.0}
        matched_pop = 0.0
        for p in regs:
            rp = region_pj[p["nuts_id"]]
            if rp:
                for k in sum_matched_pj:
                    sum_matched_pj[k] += rp[k]
                matched_pop += pop.get(p["nuts_id"], 0.0)

        raw = {}
        for p in regs:
            code = p["nuts_id"]
            rp = region_pj[code]
            rpop = pop.get(code, 0.0)
            w = {}
            for k in ("ag", "forestry", "manure"):
                if sum_matched_pj[k] <= 0:
                    # country has no ENSPRESO signal for this stream -> weight by population
                    w[k] = rpop
                elif rp is not None:
                    w[k] = rp[k]
                elif matched_pop > 0:
                    w[k] = sum_matched_pj[k] * rpop / matched_pop   # impute by per-capita
                else:
                    w[k] = rpop
            w["pop"] = rpop
            # item 11: for the UK, drive manure by housed-livestock units and ag by arable land
            # from the UK June Census (Eurostat FSS 2016), not ENSPRESO/population. Overrides the
            # weights BEFORE the per-country normalisation, so UK country totals stay exact.
            if iso3 == "GBR" and code in uk_weights:
                w["manure"] = uk_weights[code]["manure_w"]
                w["ag"] = uk_weights[code]["ag_w"]
                w["_uk_census"] = True
            raw[code] = w

        # normalization factors so each stream sums to the country total
        def factor(stream_key, total):
            s = sum(raw[p["nuts_id"]][stream_key] for p in regs)
            return (total / s) if s > 0 else 0.0
        f_ag = factor("ag", tot["ag"])
        f_for = factor("forestry", tot["forestry"])
        f_man = factor("manure", tot["manure"])
        f_msw = factor("pop", tot["msw"])
        f_wwtp = factor("pop", tot["wwtp"])

        for p in regs:
            code = p["nuts_id"]
            rp = region_pj[code]
            w = raw[code]
            ag = w["ag"] * f_ag
            forestry = w["forestry"] * f_for
            manure = w["manure"] * f_man
            msw = w["pop"] * f_msw
            wwtp = w["pop"] * f_wwtp
            biofrac = tot["biofrac"]
            dom = dominant(ag, forestry, manure, msw * biofrac)

            enspreso_src = ("JRC ENSPRESO NUTS-2 (ENS_Med 2020) share, scaled to country total"
                            if rp else "population-allocated country total (no ENSPRESO NUTS-2 match)")
            uk_census = w.get("_uk_census")
            man_src = ("UK June Agricultural Census (Eurostat FSS 2016) — housed-livestock units "
                       "(bovine+swine+poultry), scaled to country total") if uk_census else enspreso_src
            ag_src = ("UK June Agricultural Census (Eurostat FSS 2016) — arable land area, scaled "
                      "to country total") if uk_census else enspreso_src
            records.append({
                "id": p["id"],
                "name": p["name"],
                "level": "nuts2",
                "parent": iso3,                 # ISO3 -> engine manure_ad_preferred / region_country
                "cntr": p["cntr"],
                "nuts_id": code,
                "area_km2": p["area_km2"],
                "centroid": p["centroid"],
                "ag_residues_odt_mt": est(ag, source=ag_src),
                "forestry_residues_odt_mt": est(forestry, source=enspreso_src),
                "msw_total_mt": est(msw, source="country MSW total allocated by NUTS-2 population (Eurostat)"),
                "msw_biogenic_frac": {"value": biofrac, "source": "country value (global tool)"},
                "animal_manure_odt_mt": est(manure, source=man_src,
                                            notes=("housed cattle/pig/poultry manure" if uk_census
                                                   else "ENSPRESO biogas feedstock (manure_liq+sol)")),
                "human_wwtp_odt_mt": est(wwtp, source="country biosolids total allocated by NUTS-2 population"),
                "nutrient_status": tot["nutrient_status"],
                "nutrient_status_source": "inherited from country (global tool)",
                "dominant_feedstock": dom,
                "feedstock_density": "diffuse",   # engine recomputes from area
                "notes": "NUTS-2 disaggregation; see METHODOLOGY.md §EU",
            })

    records.sort(key=lambda r: r["nuts_id"])
    with open(OUT, "w") as f:
        json.dump(records, f, ensure_ascii=False)

    def tot_stream(k):
        return sum(r[k]["value"] for r in records)
    from collections import Counter
    print(f"wrote {len(records)} NUTS-2 feedstock records -> {OUT}")
    print(f"  ag {tot_stream('ag_residues_odt_mt'):.1f}  forestry {tot_stream('forestry_residues_odt_mt'):.1f}"
          f"  manure {tot_stream('animal_manure_odt_mt'):.1f}  msw {tot_stream('msw_total_mt'):.1f}  (Mt)")
    print(f"  dominant: {dict(Counter(r['dominant_feedstock'] for r in records))}")
    print(f"  population-fallback regions (no ENSPRESO match): {len(fallbacks)}")


if __name__ == "__main__":
    main()
