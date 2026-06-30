#!/usr/bin/env python3
"""
US infrastructure layers for the county BiCRS map:

  data/processed/facilities_us_detailed.json
    Granular biogenic point sources + large WWTPs, from EPA GHGRP 2023 (Direct Point
    Emitters): facility coordinates + reported biogenic CO2. Types via NAICS:
      pulp_paper (322x) · wte (562213) · landfill (562212) · ethanol (311221/3251x) ·
      bioenergy (2211x/321x/611310) · wwtp (221320, sized by total reported emissions).
    A facility is kept if biogenic CO2 >= 25 kt/yr, OR it is biomass-dominated
    (biogenic >= 50% of reported CO2), OR it is a reporting sewage plant.

  data/processed/wells_us.json
    CO2 storage / injection points:
      - Operational geologic sequestration (GHGRP Subpart RR reporters: real MRV sites).
      - Class VI wells (issued / draft / pending) — curated from the EPA Class VI Data
        Repository (current to 2026).
      - Class V projects — curated biomass-injection / bio-oil sequestration (Vaulted, Charm).

Input: data/geo/us_raw/ghgrp/ghgp_data_2023.xlsx
"""
import json
import os

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
PROC = os.path.join(ROOT, "data", "processed")
RAW = os.path.join(ROOT, "data", "geo", "us_raw")
XLSX = os.path.join(RAW, "ghgrp", "ghgp_data_2023.xlsx")
WWTP_RAW = os.path.join(RAW, "wwtp_major.ndjson")
LMOP_XLSX = os.path.join(RAW, "lmop", "lmopcompositedata.xlsx")

FAC_OUT = os.path.join(PROC, "facilities_us_detailed.json")
WELLS_OUT = os.path.join(PROC, "wells_us.json")
WWTP_OUT = os.path.join(PROC, "wwtps_us.json")

COUNTIES = os.path.join(ROOT, "data", "geo", "us_counties.json")
AD_XLSX = os.path.join(ROOT, "data", "geo", "ad_raw", "agstar.xlsx")

BIO_MIN = 25000.0  # metric tons biogenic CO2/yr threshold for point sources
# Landfill gas (LMOP): collected-gas biogenic CO2 if fully combusted. 1 lb-mol of ideal gas =
# 379.48 scf (60 degF, 14.696 psia); LFG is ~CH4 + CO2 (~96%), remainder inert. Combustion turns
# CH4 -> CO2 and the raw CO2 passes through, so CO2 yield = (CH4 frac + CO2 frac) of the gas.
LBMOL_SCF = 379.48
LFG_MIN_CO2_MTPA = 0.05    # qualifying-landfill floor (matches engine_core.LFG_MIN_CO2_MTPA)
# Capturable biogenic CO2 per (cu-ft/day) of biogas, Mtpa: ~total biogas carbon -> CO2
# (biogas ~60% CH4 / 40% CO2; ~1.98 kg CO2 per m3 biogas).
CFD_TO_MTPA = 0.0283168 * 365 * 1.98 / 1e9
AD_DEFAULT_MTPA = 0.003   # per digester when biogas estimate is missing


def naics_type(code):
    c = str(code)
    if c.startswith("322"):
        return "pulp_paper"
    if c == "562213":
        return "wte"
    if c == "562212":
        return "landfill"
    if c == "311221" or c.startswith("3251"):
        return "ethanol"
    if c == "221320":
        return "wwtp"
    if c.startswith("2211") or c.startswith("321") or c == "611310" or c == "221330":
        return "bioenergy"
    return None


def fnum(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def title_case(s):
    s = (s or "").strip()
    # Many GHGRP names are ALL CAPS; title-case those, leave mixed-case alone.
    return s.title() if s.isupper() else s


def retrofit_score(ftype, bio_mt):
    if ftype in ("pulp_paper", "ethanol", "wte"):
        return "high"
    if ftype == "bioenergy":
        return "high" if bio_mt >= 0.2 else "medium"
    if ftype == "wwtp":
        return "medium"
    return "low"  # landfill — dilute biogenic, lower retrofit attractiveness


def build_facilities(wb):
    ws = wb["Direct Point Emitters"]
    it = ws.iter_rows(values_only=True)
    for _ in range(4):
        next(it)
    facs = []
    for r in it:
        name, state = r[2], r[4]
        lat, lon = fnum(r[8]), fnum(r[9])
        naics = r[10]
        total_nonbio = fnum(r[13]) or 0.0
        bio = fnum(r[25]) or 0.0
        if lat is None or lon is None:
            continue
        ftype = naics_type(naics)
        if ftype is None or ftype == "wwtp":
            continue  # WWTPs handled in the dedicated Major-POTW layer (build_wwtps)
        if ftype == "landfill":
            continue  # landfills handled from EPA LMOP (gas-collection + LFG flow) in build_landfills

        biofrac = bio / (bio + total_nonbio) if (bio + total_nonbio) > 0 else 0
        if bio < BIO_MIN and biofrac < 0.5:
            continue
        size_mt = round(bio / 1e6, 4)
        note = f"{int(bio):,} t biogenic CO2/yr (GHGRP 2023)"

        facs.append({
            "name": title_case(name),
            "type": ftype,
            "country": "USA",
            "state": state,
            "lat": round(lat, 4),
            "lon": round(lon, 4),
            "capacity_note": note,
            "est_biogenic_co2_mtpa": {"value": size_mt},
            "retrofit_score": retrofit_score(ftype, size_mt),
            "existing": True,
            "source": "EPA GHGRP 2023 (Direct Point Emitters)",
            "naics": str(naics),
        })
    facs.sort(key=lambda f: -f["est_biogenic_co2_mtpa"]["value"])
    return facs


def _lmop_co2_mtpa(mmscfd, ch4frac):
    """Collected landfill-gas biogenic CO2 if fully combusted (Mt/yr)."""
    if not mmscfd:
        return 0.0
    co2frac = max(0.0, 0.96 - ch4frac)
    return mmscfd * 1e6 * 365 / LBMOL_SCF * 44 / 2204.6 * (ch4frac + co2frac) / 1e6


def build_landfills():
    """EPA LMOP gas-collecting landfills as LFG+CCS / LFG-RNG+CCS retrofit anchors.

    Keeps landfills with a gas-collection system in place, coordinates, and collected-gas biogenic
    CO2 (full-combustion basis) >= LFG_MIN_CO2_MTPA. Tags each with whether it runs/plans an RNG
    upgrading project (-> LFG-RNG+CCS is the natural retrofit: the upgrading already vents a near-pure
    CO2 stream) vs electricity/flare (-> LFG combustion+CCS). This is the canonical LFG dataset —
    location, waste-in-place, LFG collected (mmscfd), gas-collection Y/N, %CH4, project type — far
    richer than the GHGRP landfill rows it replaces."""
    if not os.path.exists(LMOP_XLSX):
        print("  (LMOP file missing; skipping landfills)")
        return []
    wb = openpyxl.load_workbook(LMOP_XLSX, read_only=True)
    ws = wb["LMOP Database"]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    H = {h: i for i, h in enumerate(hdr)}

    def g(r, k):
        return r[H[k]] if k in H else None

    lfs = {}   # Landfill ID -> {row, rng, elec} (a landfill can have multiple project rows)
    for r in it:
        lid = g(r, "Landfill ID")
        if lid is None:
            continue
        d = lfs.setdefault(lid, {"row": r, "rng": False, "elec": False})
        cat = str(g(r, "Project Type Category") or "")
        status = str(g(r, "Current Project Status") or "").lower()
        live = any(s in status for s in ("operational", "construction", "planned", "design"))
        if "Renewable Natural Gas" in cat and live:
            d["rng"] = True
        if "Electricity" in cat and live:
            d["elec"] = True

    out = []
    for d in lfs.values():
        r = d["row"]
        gc = str(g(r, "LFG Collection System In Place?") or "").strip().lower() in ("yes", "y")
        lat, lon = fnum(g(r, "Latitude")), fnum(g(r, "Longitude"))
        if not (gc and lat and lon):
            continue
        coll = fnum(g(r, "LFG Collected (mmscfd)")) or fnum(g(r, "LFG Generated (mmscfd)"))
        pm = fnum(g(r, "Percent Methane"))
        ch4 = pm / 100.0 if pm else 0.5
        co2 = _lmop_co2_mtpa(coll, ch4)
        if co2 < LFG_MIN_CO2_MTPA:
            continue
        pref = "rng" if d["rng"] else "ccs"
        proj = ("gas-to-RNG project" if d["rng"] else
                "electricity project" if d["elec"] else "flare / collection only")
        out.append({
            "name": title_case(str(g(r, "Landfill Name") or "")),
            "type": "landfill",
            "country": "USA",
            "state": g(r, "State"),
            "lat": round(lat, 4),
            "lon": round(lon, 4),
            "capacity_note": f"{co2 * 1e6:,.0f} t collected-gas biogenic CO2/yr "
                             f"(LFG, full-combustion basis); {proj}",
            "est_biogenic_co2_mtpa": {"value": round(co2, 4)},
            "gas_collection": True,
            "lfg_project": pref,          # "rng" -> LFG-RNG+CCS preferred; "ccs" -> LFG combustion+CCS
            "retrofit_score": "high" if co2 >= 0.15 else "medium",
            "existing": True,
            "source": "EPA LMOP Landfill & Project Database (2024)",
            "naics": "562212",
        })
    out.sort(key=lambda f: -f["est_biogenic_co2_mtpa"]["value"])
    nrng = sum(1 for f in out if f["lfg_project"] == "rng")
    print(f"  US landfills (LMOP, gas-collection, >= {LFG_MIN_CO2_MTPA} Mt CO2/yr): {len(out)} "
          f"({nrng} with an RNG project)")
    return out


def build_sequestration_wells(wb):
    """Operational geologic-sequestration reporters (GHGRP Subpart RR) — real MRV sites."""
    ws = wb["Geologic Sequestration of CO2"]
    it = ws.iter_rows(values_only=True)
    for _ in range(4):
        next(it)
    wells = []
    for r in it:
        name = title_case(r[2])
        lat, lon = fnum(r[8]), fnum(r[9])
        stored = fnum(r[12]) or 0.0
        if lat is None or lon is None:
            continue
        wells.append({
            "name": name,
            "kind": "sequestration",     # operational geologic sequestration (Subpart RR)
            "well_class": "VI/RR",
            "status": "operational",
            "state": r[4],
            "lat": round(lat, 4),
            "lon": round(lon, 4),
            "co2_mtpa": round(stored / 1e6, 4),
            "source": "EPA GHGRP 2023 (Subpart RR geologic sequestration)",
        })
    wells.sort(key=lambda w: -w["co2_mtpa"])
    return wells


# --- Curated Class VI wells (EPA Class VI Data Repository, current to 2026) and Class V
#     biomass-injection / bio-oil projects. Coordinates are project-site approximate. ---
CURATED_WELLS = [
    # ----- Class VI: ISSUED permits -----
    {"name": "ADM Decatur (CCS#1/#2)", "well_class": "VI", "status": "issued",
     "state": "IL", "lat": 39.8675, "lon": -88.885, "operator": "Archer Daniels Midland",
     "source": "EPA Class VI permit (Decatur, IL)"},
    {"name": "Red Trail Energy", "well_class": "VI", "status": "issued",
     "state": "ND", "lat": 46.879, "lon": -102.318, "operator": "Red Trail Energy",
     "source": "ND DMR Class VI (primacy)"},
    {"name": "Blue Flint / Harvestone", "well_class": "VI", "status": "issued",
     "state": "ND", "lat": 47.34, "lon": -101.43, "operator": "Harvestone Low Carbon Partners",
     "source": "ND DMR Class VI (primacy)"},
    {"name": "Minnkota Project Tundra", "well_class": "VI", "status": "issued",
     "state": "ND", "lat": 47.36, "lon": -101.86, "operator": "Minnkota Power",
     "source": "ND DMR Class VI (primacy)"},
    {"name": "PureField Carbon Capture", "well_class": "VI", "status": "issued",
     "state": "KS", "lat": 38.89, "lon": -98.86, "operator": "PureField Ingredients",
     "source": "EPA Region 7 Class VI permit (2026)"},
    {"name": "Wabash Carbon Services", "well_class": "VI", "status": "issued",
     "state": "IN", "lat": 39.53, "lon": -87.43, "operator": "Wabash Valley Resources",
     "source": "EPA Region 5 Class VI permit"},
    # ----- Class VI: DRAFT / proposed (representative, 2026) -----
    {"name": "Hackberry Carbon Sequestration", "well_class": "VI", "status": "draft",
     "state": "LA", "lat": 29.99, "lon": -93.34, "operator": "Sempra / Hackberry",
     "source": "EPA/LA Class VI draft permit"},
    {"name": "CapturePoint Central Louisiana", "well_class": "VI", "status": "draft",
     "state": "LA", "lat": 31.3, "lon": -92.4, "operator": "CapturePoint Solutions",
     "source": "LA Class VI (primacy) draft"},
    {"name": "Bayou Bend CCS", "well_class": "VI", "status": "pending",
     "state": "TX", "lat": 29.76, "lon": -93.95, "operator": "Chevron/Talos/Equinor",
     "source": "EPA Region 6 Class VI application"},
    {"name": "Tenaska Bison Lasso", "well_class": "VI", "status": "pending",
     "state": "TX", "lat": 32.5, "lon": -97.0, "operator": "Tenaska",
     "source": "EPA Region 6 Class VI application"},
    {"name": "Heartland Greenway (Navigator)", "well_class": "VI", "status": "pending",
     "state": "IL", "lat": 39.9, "lon": -89.6, "operator": "Navigator CO2",
     "source": "EPA Region 5 Class VI application"},
    {"name": "Wyoming Frontier CarbonSAFE", "well_class": "VI", "status": "pending",
     "state": "WY", "lat": 41.79, "lon": -107.2, "operator": "Frontier/CarbonSAFE",
     "source": "WY DEQ Class VI (primacy)"},
    # ----- Class V: biomass injection / bio-oil sequestration -----
    {"name": "Vaulted Deep (Kansas)", "well_class": "V", "status": "operational",
     "state": "KS", "lat": 37.7, "lon": -97.9, "operator": "Vaulted Deep",
     "source": "Vaulted Deep — biomass/organic-waste slurry injection (Class V)"},
    {"name": "Vaulted Deep (Los Angeles)", "well_class": "V", "status": "operational",
     "state": "CA", "lat": 33.9, "lon": -118.2, "operator": "Vaulted Deep",
     "source": "Vaulted Deep — biosolids slurry injection (Class V)"},
    {"name": "Charm Industrial (Permian)", "well_class": "V", "status": "operational",
     "state": "TX", "lat": 31.9, "lon": -102.1, "operator": "Charm Industrial",
     "source": "Charm Industrial — bio-oil injection into EOR/saltwater-disposal wells"},
    {"name": "Charm Industrial (Central Valley)", "well_class": "V", "status": "operational",
     "state": "CA", "lat": 36.3, "lon": -119.8, "operator": "Charm Industrial",
     "source": "Charm Industrial — bio-oil injection"},
]


def build_wwtps():
    """Major NPDES POTWs (design flow >= 1 MGD) from EPA FRS — large WWTP point sources."""
    out = []
    if not os.path.exists(WWTP_RAW):
        return out
    with open(WWTP_RAW) as f:
        for line in f:
            a = json.loads(line)
            lat, lon = fnum(a.get("FAC_LAT")), fnum(a.get("FAC_LONG"))
            if lat is None or lon is None:
                continue
            out.append({
                "name": title_case(a.get("CWP_NAME") or ""),
                "type": "wwtp",
                "state": a.get("CWP_STATE"),
                "lat": round(lat, 4),
                "lon": round(lon, 4),
                "source": "EPA FRS / NPDES — Major POTW (>=1 MGD)",
            })
    return out


def _norm_county(s):
    s = (s or "").lower()
    for suf in (" county", " parish", " borough", " census area", " municipality", " city and borough"):
        if s.endswith(suf):
            s = s[: -len(suf)]
    s = s.replace("saint ", "st ").replace("ste ", "st ")
    return "".join(ch for ch in s if ch.isalnum())


def build_ad(counties):
    """EPA AgSTAR livestock digesters -> per-county cumulative AD nodes (biogas_ad).

    AgSTAR gives city/county/state (no lat/lon); we map each operational/construction
    digester to its county centroid (us_counties.json) and aggregate capturable biogenic CO2
    per county. ADs are individually small, so the county node carries the *cumulative*
    capacity — what matters for an AD+CCS retrofit gate. Coverage radius = 15 km (wet-manure
    haul is short)."""
    if not os.path.exists(AD_XLSX):
        print("  (AgSTAR file missing; skipping US AD)")
        return []
    # (state, normalized county) -> {centroid, fips, name}
    idx = {}
    for f in counties:
        p = f["properties"]
        idx[(p["state"], _norm_county(p["name"]))] = p
    wb = openpyxl.load_workbook(AD_XLSX, read_only=True)
    ws = wb["Operational and Construction"]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    ci = {h: i for i, h in enumerate(hdr)}
    i_county, i_state = ci["County"], ci["State"]
    i_biogas = ci.get("Biogas Generation Estimate (cu-ft/day)")
    agg = {}   # fips -> {co2, n, props}
    matched = unmatched = 0
    for r in it:
        if not r or r[0] is None:
            continue
        st, cty = r[i_state], r[i_county]
        prop = idx.get((st, _norm_county(cty)))
        if not prop:
            unmatched += 1
            continue
        matched += 1
        co2 = AD_DEFAULT_MTPA
        bg = fnum(r[i_biogas]) if i_biogas is not None else None
        if bg:
            co2 = bg * CFD_TO_MTPA
        fips = prop["fips"]
        a = agg.setdefault(fips, {"co2": 0.0, "n": 0, "p": prop})
        a["co2"] += co2
        a["n"] += 1
    out = []
    for fips, a in agg.items():
        p = a["p"]
        out.append({
            "name": f"{p['name']} County AD ({a['n']} digester{'s' if a['n'] != 1 else ''})",
            "type": "biogas_ad", "country": "USA", "state": p["state"],
            "lat": round(p["centroid"][1], 4), "lon": round(p["centroid"][0], 4),
            "capacity_note": f"{a['n']} livestock digester(s), cumulative",
            "est_biogenic_co2_mtpa": {"value": round(a["co2"], 4)},
            "retrofit_score": "medium", "existing": True, "proc_radius_km": 15,
            "source": "EPA AgSTAR Livestock Anaerobic Digester Database (aggregated to county)",
        })
    print(f"  US AD: {matched} digesters matched, {unmatched} unmatched -> {len(out)} county AD nodes")
    return out


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    facs = build_facilities(wb)
    counties = json.load(open(COUNTIES))["features"]
    facs += build_ad(counties)
    facs += build_landfills()
    seq = build_sequestration_wells(wb)
    wwtps = build_wwtps()

    # Merge operational sequestration + curated Class VI/V into one wells layer.
    wells = list(seq)
    for w in CURATED_WELLS:
        w = dict(w)
        w.setdefault("kind", "class6" if w["well_class"] in ("VI", "VI/RR") else "class5")
        w.setdefault("co2_mtpa", None)
        wells.append(w)

    with open(FAC_OUT, "w") as f:
        json.dump(facs, f, ensure_ascii=False, indent=1)
    with open(WELLS_OUT, "w") as f:
        json.dump(wells, f, ensure_ascii=False, indent=1)
    with open(WWTP_OUT, "w") as f:
        json.dump(wwtps, f, ensure_ascii=False)

    from collections import Counter
    print(f"facilities: {len(facs)}  by type {dict(Counter(x['type'] for x in facs))}")
    print(f"wells: {len(wells)}  "
          f"(RR operational {len(seq)}, curated {len(CURATED_WELLS)})  "
          f"by class {dict(Counter(x['well_class'] for x in wells))}")
    print(f"wwtps (Major POTWs): {len(wwtps)}")


if __name__ == "__main__":
    main()
